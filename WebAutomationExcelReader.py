from openpyxl import load_workbook


class WebAutomationExcelReader:
    """Custom Autosphere library to read challenge.xlsx"""

    def __init__(self):
        self._wb = None
        self._ws = None

    def read_excel_file(self, filepath):
        """Open an Excel file and load its active worksheet.
        
        Args:
            filepath (str): Full path to the .xlsx file
            
        Returns:
            str: Summary of rows read (e.g. 'Read 50 data rows from sheet: data')
        """
        self._wb = load_workbook(filepath, read_only=True)
        self._ws = self._wb.active
        row_count = self._ws.max_row - 1  # exclude header
        sheet_name = self._ws.title
        return f"Read {row_count} data rows from sheet: {sheet_name}"

    def get_excel_data_as_list(self):
        """Return all worksheet rows as a list of dictionaries.
        
        First row is used as header / key names.
        
        Returns:
            list: List of dicts, one per data row
        """
        if self._ws is None:
            raise RuntimeError("No Excel file open. Call 'Read Excel File' first.")
        
        rows = list(self._ws.iter_rows(values_only=True))
        headers = [str(col) for col in rows[0]]
        result = []
        for row in rows[1:]:
            result.append(dict(zip(headers, row)))
        return result
